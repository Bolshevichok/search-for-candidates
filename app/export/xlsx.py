from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db.repository import Repository
from app.pipeline.cancellation import CancellationToken
from app.sources.vak.parser import split_specialty

SITE_STATUSES = frozenset({"site_no_vak", "site_and_vak", "site_and_vak_probable"})
VAK_STATUSES = frozenset({"vak_no_site", "site_and_vak", "site_and_vak_probable"})

SITE_COLUMNS = [
  "full_name",
  "match_status",
  "university",
  "post",
  "degree",
  "academic_title",
  "disciplines",
  "spec_experience",
  "email",
  "phone",
  "vk_profile",
  "vk_email",
  "vk_phone",
  "source_url",
]

SITE_HEADERS = [
  "ФИО",
  "Где найден кандидат",
  "Университет",
  "Должность",
  "Учёная степень",
  "Учёное звание",
  "Преподаваемые дисциплины",
  "Стаж по специальности, лет",
  "Электронная почта",
  "Телефон",
  "VK профиль",
  "VK: публичный e-mail",
  "VK: публичный телефон",
  "Страница источника",
]

VAK_COLUMNS = [
  "full_name",
  "match_status",
  "branch",
  "specialty_code",
  "specialty_name",
  "dissertation_type",
  "topic",
  "defense_date",
  "defend_org",
  "council_cipher",
  "org_address",
  "org_phone",
  "is_pilot",
  "vk_profile",
  "vk_email",
  "vk_phone",
]

VAK_HEADERS = [
  "ФИО",
  "Где найден кандидат",
  "Отрасль науки",
  "Код специальности",
  "Название специальности",
  "Тип диссертации",
  "Тема диссертации",
  "Дата защиты",
  "Организация защиты",
  "Шифр совета",
  "Адрес организации",
  "Телефон организации",
  "Порядок присуждения степени",
  "VK профиль",
  "VK: публичный e-mail",
  "VK: публичный телефон",
]

MATCH_STATUS_LABELS = {
  "site_and_vak": "Найден на сайте вуза и в ВАК",
  "site_and_vak_probable": "Вероятно найден на сайте вуза и в ВАК",
  "vak_no_site": "Найден в ВАК, не найден на сайте вуза",
  "site_no_vak": "Найден на сайте вуза, не найден в ВАК",
}

HEADER_NOTES = {
  "Где найден кандидат": (
    "Показывает, в каком источнике найдена запись: на сайте вуза, в базе ВАК или в обоих."
  ),
  "Порядок присуждения степени": (
    "Самостоятельное присуждение — степень присуждена вузом или научной организацией, "
    "которые имеют такое право. Защита через диссертационный совет ВАК — обычный порядок защиты."
  ),
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
_MAX_COLUMN_WIDTH = 50
_MIN_COLUMN_WIDTH = 12


def _format_sheet(ws: Any) -> None:
  """Make an exported sheet readable without manual Excel adjustments."""
  ws.freeze_panes = "A2"
  ws.auto_filter.ref = ws.dimensions
  ws.row_dimensions[1].height = 32

  for cell in ws[1]:
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = _HEADER_ALIGNMENT
    if cell.value in HEADER_NOTES:
      cell.comment = Comment(HEADER_NOTES[cell.value], "Поиск кандидатов")

  for row in ws.iter_rows(min_row=2):
    for cell in row:
      cell.alignment = _CELL_ALIGNMENT

  for column_cells in ws.iter_cols():
    width = max(
      (len(str(cell.value)) for cell in column_cells if cell.value is not None),
      default=0,
    )
    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
      max(width + 2, _MIN_COLUMN_WIDTH), _MAX_COLUMN_WIDTH
    )


def _latest_defense(defenses_json: str | None) -> dict[str, Any]:
  if not defenses_json:
    return {}
  items = json.loads(defenses_json)
  if not items:
    return {}
  return max(items, key=lambda d: d.get("date") or "")


def _format_duration(started_at: str, finished_at: str) -> str:
  seconds = max(0, int((datetime.fromisoformat(finished_at) - datetime.fromisoformat(started_at)).total_seconds()))
  hours, seconds = divmod(seconds, 3600)
  minutes, seconds = divmod(seconds, 60)
  return f"{hours:02}:{minutes:02}:{seconds:02}"


def _match_status_label(status: str) -> str:
  return MATCH_STATUS_LABELS.get(status, status)


def _degree_award_process(value: Any) -> str:
  if value:
    return "Самостоятельное присуждение степени вузом"
  return "Защита через диссертационный совет ВАК"


def _defense_fields(defense: dict[str, Any]) -> dict[str, Any]:
  specialty = defense.get("specialty")
  code = defense.get("specialty_code") or ""
  name = defense.get("specialty_name") or ""
  if specialty and not code and not name:
    code, name = split_specialty(specialty)
  return {
    "branch": defense.get("branch") or "",
    "specialty_code": code,
    "specialty_name": name,
    "dissertation_type": defense.get("dissertation_type") or "",
    "topic": defense.get("topic") or "",
    "defense_date": defense.get("date") or "",
    "defend_org": defense.get("defend_org") or "",
    "council_cipher": defense.get("council_cipher") or "",
    "org_address": defense.get("org_address") or "",
    "org_phone": defense.get("org_phone") or "",
    "is_pilot": defense.get("is_pilot") if defense else "",
  }


def _best_vk_profile(repo: Repository, candidate_id: str) -> dict[str, Any]:
  rows = repo.execute(
    """
    SELECT p.*, vc.vk_url AS community_url
    FROM candidate_vk_profiles p
    JOIN university_vk_communities vc ON vc.community_id = p.community_id
    WHERE p.candidate_id = ? AND p.profile_url <> ''
    ORDER BY CASE p.vk_match_status
      WHEN 'matched' THEN 0
      WHEN 'ambiguous' THEN 1
      WHEN 'not_found' THEN 2
      ELSE 3 END,
      p.checked_at DESC
    """,
    (candidate_id,),
  ).fetchall()
  if not rows:
    return {}
  profile = dict(rows[0])
  profile["profile_url"] = "\n".join(dict.fromkeys(row["profile_url"] for row in rows))
  return profile


def export_xlsx(
  repo: Repository,
  output_path: Path,
  domain: str | None = None,
  *,
  cancel_token: CancellationToken | None = None,
) -> Path:
  if cancel_token is not None:
    cancel_token.check()
  output_path.parent.mkdir(parents=True, exist_ok=True)
  partial_path = output_path.with_name(f".{output_path.name}.partial")
  wb = Workbook()

  ws_site = wb.active
  ws_site.title = "Сотрудники вузов"
  ws_site.append(SITE_HEADERS)

  ws_vak = wb.create_sheet("Кандидаты ВАК")
  ws_vak.append(VAK_HEADERS)

  universities = {
    int(row["university_id"]): row
    for row in repo.execute("SELECT * FROM universities").fetchall()
  }
  allowed_university_ids: set[int] | None = None
  if domain is not None:
    allowed_university_ids = {
      uid for uid, row in universities.items() if row["domain"] == domain
    }

  for row in repo.execute("SELECT * FROM candidates ORDER BY candidate_id").fetchall():
    if cancel_token is not None:
      cancel_token.check()
    if allowed_university_ids is not None and int(row["university_id"] or -1) not in allowed_university_ids:
      continue
    status = row["match_status"]
    uni = universities.get(int(row["university_id"])) if row["university_id"] else None
    disciplines = json.loads(row["disciplines"] or "[]") if row["disciplines"] else []
    defense = _latest_defense(row["defenses"])
    vak_fields = _defense_fields(defense)
    vk_profile = _best_vk_profile(repo, row["candidate_id"])

    if status in SITE_STATUSES:
      ws_site.append(
        [
          row["full_name"],
          _match_status_label(status),
          uni["official_name"] if uni else "",
          row["post"] or "",
          row["degree"] or "",
          row["academic_title"] or "",
          "; ".join(disciplines),
          row["spec_experience"] if row["spec_experience"] is not None else "",
          row["email"] or "",
          row["phone"] or "",
          vk_profile.get("profile_url") or "",
          vk_profile.get("public_email") or "",
          vk_profile.get("public_phone") or "",
          row["source_url"] or "",
        ]
      )

    if status in VAK_STATUSES:
      ws_vak.append(
        [
          row["full_name"],
          _match_status_label(status),
          vak_fields["branch"],
          vak_fields["specialty_code"],
          vak_fields["specialty_name"],
          vak_fields["dissertation_type"],
          vak_fields["topic"],
          vak_fields["defense_date"],
          vak_fields["defend_org"],
          vak_fields["council_cipher"],
          vak_fields["org_address"],
          vak_fields["org_phone"],
          _degree_award_process(vak_fields["is_pilot"]),
          vk_profile.get("profile_url") or "",
          vk_profile.get("public_email") or "",
          vk_profile.get("public_phone") or "",
        ]
      )

  ws_namesakes = wb.create_sheet("Возможные тёзки")
  ws_namesakes.append(
    [
      "ФИО на сайте вуза",
      "Университет",
      "ФИО в базе ВАК",
      "Организация защиты",
      "Причина проверки",
    ]
  )
  for row in repo.execute(
    """
    SELECT pn.reason,
           sc.full_name AS site_full_name,
           sv.official_name AS site_university,
           vc.full_name AS vak_full_name,
           json_extract(vc.defenses, '$[0].defend_org') AS vak_defend_org
    FROM possible_namesakes pn
    JOIN candidates sc ON sc.candidate_id = pn.site_candidate_id
    JOIN candidates vc ON vc.candidate_id = pn.vak_candidate_id
    LEFT JOIN universities sv ON sv.university_id = sc.university_id
  """
  ).fetchall():
    if cancel_token is not None:
      cancel_token.check()
    ws_namesakes.append(
      [
        row["site_full_name"],
        row["site_university"] or "",
        row["vak_full_name"],
        row["vak_defend_org"] or "",
        row["reason"],
      ]
    )

  ws_errors = wb.create_sheet("Ошибки вузов")
  ws_errors.append(["Университет", "Домен", "Тип ошибки", "Последняя попытка"])
  for row in repo.execute(
    """
    SELECT u.official_name, u.domain, ue.error_type, ue.last_attempt_at
    FROM university_errors ue
    JOIN universities u ON u.university_id = ue.university_id
    ORDER BY ue.id
  """
  ).fetchall():
    if cancel_token is not None:
      cancel_token.check()
    ws_errors.append(
      [row["official_name"], row["domain"] or "", row["error_type"], row["last_attempt_at"]]
    )

  ws_meta = wb.create_sheet("Сведения о выгрузке")
  ws_meta.append(
    [
      "ID запуска",
      "Область выгрузки",
      "Начало",
      "Окончание",
      "Длительность",
      "Вузов обработано успешно",
      "Вузов с ошибками",
      "Всего кандидатов в выгрузке",
      "Найдено VK-профилей в выгрузке",
      "Есть на сайте и в ВАК в выгрузке",
      "Вероятно есть на сайте и в ВАК в выгрузке",
      "Есть в ВАК, нет на сайте в выгрузке",
      "Есть на сайте, нет в ВАК в выгрузке",
    ]
  )
  run = repo.execute("SELECT * FROM runs ORDER BY run_id DESC LIMIT 1").fetchone()
  if run:
    finished_at = run["finished_at"] or datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    scope_params: list[Any] = []
    if domain is not None:
      scope_params.append(domain)
    ok = repo.execute(
      f"SELECT COUNT(*) AS c FROM universities WHERE layer1_status = 'ok'"
      + (" AND domain = ?" if domain is not None else ""),
      scope_params,
    ).fetchone()["c"]
    err_sql = """
      SELECT COUNT(*) AS c
      FROM university_errors ue
      JOIN universities u ON u.university_id = ue.university_id
      WHERE ue.run_id = ?
    """
    err_params: list[Any] = [run["run_id"]]
    if domain is not None:
      err_sql += " AND u.domain = ?"
      err_params.append(domain)
    err = repo.execute(err_sql, err_params).fetchone()["c"]
    counts_sql = """
      SELECT c.match_status, COUNT(*) AS c
      FROM candidates c
      LEFT JOIN universities u ON u.university_id = c.university_id
    """
    count_params: list[Any] = []
    if domain is not None:
      counts_sql += " WHERE u.domain = ?"
      count_params.append(domain)
    counts_sql += " GROUP BY c.match_status"
    counts = {
      row["match_status"]: row["c"]
      for row in repo.execute(counts_sql, count_params).fetchall()
    }
    total = sum(counts.values())
    vk_sql = """
      SELECT COUNT(DISTINCT p.candidate_id) AS c
      FROM candidate_vk_profiles p
      JOIN candidates c ON c.candidate_id = p.candidate_id
      JOIN universities u ON u.university_id = c.university_id
      WHERE p.vk_match_status = 'matched'
    """
    vk_params: list[Any] = []
    if domain is not None:
      vk_sql += " AND u.domain = ?"
      vk_params.append(domain)
    vk_matches = repo.execute(vk_sql, vk_params).fetchone()["c"]
    ws_meta.append(
      [
        run["run_id"],
        domain or "все вузы",
        run["started_at"],
        finished_at,
        _format_duration(run["started_at"], finished_at),
        ok,
        err,
        total,
        vk_matches,
        counts.get("site_and_vak", 0),
        counts.get("site_and_vak_probable", 0),
        counts.get("vak_no_site", 0),
        counts.get("site_no_vak", 0),
      ]
    )

  for worksheet in wb.worksheets:
    if cancel_token is not None:
      cancel_token.check()
    _format_sheet(worksheet)

  wb.save(partial_path)
  if cancel_token is not None:
    cancel_token.check()
  partial_path.replace(output_path)
  return output_path


def default_output_path() -> Path:
  stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
  return Path("output") / f"candidates_{stamp}.xlsx"
